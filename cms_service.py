"""Valida el Excel CMS y sincroniza todas las bases consumidas por el proyecto."""

from __future__ import annotations

import csv
import io
import json
import os
import shutil
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from schema import (
    LOCATION_HEADERS, LOCATION_MAP, METRIC_HEADERS, METRIC_MAP,
    NUMBER_FIELDS, PERCENT_FIELDS,
)

LEGACY_HEADERS = [
    "CC Nombre", "CC", "DM", "Fecha de Apertura", "Rolling", "TPLH", "IPLH",
    "Labor", "Dif Labor%", "Ventas", "Dif ppto%", "OMT #", "NPS", "Conexión",
    "Bebida", "CTC", "Costo", "Dif costo%", "EBITDA", "Dif EBITDA %", "DT Time",
    "Tiempo DT AA", "Municipio o Delegación", "Cobertura", "Formato",
    "Formato - Comercial", "TIER", "Corte YTD", "Latitud", "Longitud",
]

LEGACY_FIELDS = {
    "CC Nombre": "store_name", "CC": "cc", "DM": "dm",
    "Fecha de Apertura": "opening_date", "Rolling": "rolling_pct", "TPLH": "tplh",
    "IPLH": "iplh", "Labor": "labor_pct", "Dif Labor%": "labor_variance_pct",
    "Ventas": "sales_mxn", "Dif ppto%": "budget_variance_pct", "OMT #": "omt",
    "NPS": "nps", "Conexión": "connection_pct", "Bebida": "beverage_pct",
    "CTC": "ctc_pct", "Costo": "cost_pct", "Dif costo%": "cost_variance_pct",
    "EBITDA": "ebitda_pct", "Dif EBITDA %": "ebitda_variance_pct",
    "DT Time": "dt_time", "Tiempo DT AA": "dt_time_aa",
    "Municipio o Delegación": "municipality", "Cobertura": "coverage",
    "Formato": "format", "Formato - Comercial": "commercial_format", "TIER": "tier",
    "Corte YTD": "cutoff_ytd", "Latitud": "latitude", "Longitud": "longitude",
}


class CMSValidationError(ValueError):
    def __init__(self, errors: list[str]):
        super().__init__("; ".join(errors))
        self.errors = errors


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def _serialize(value: Any) -> Any:
    value = _clean(value)
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return value


def _read_sheet(ws, expected_headers: list[str]) -> list[dict[str, Any]]:
    headers = [_clean(cell.value) for cell in ws[1]]
    missing = [header for header in expected_headers if header not in headers]
    if missing:
        raise CMSValidationError([f"{ws.title}: faltan columnas: {', '.join(missing)}"])
    positions = {header: headers.index(header) for header in expected_headers}
    rows = []
    for row_number, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        record = {header: _serialize(cells[index] if index < len(cells) else None) for header, index in positions.items()}
        if any(value is not None for value in record.values()):
            record["_row"] = row_number
            rows.append(record)
    return rows


def _to_float(value: Any, label: str, errors: list[str], allow_none: bool = True) -> float | None:
    value = _clean(value)
    if value is None and allow_none:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        errors.append(f"{label}: debe ser numérico")
        return None


def workbook_to_payload(path: str | Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    errors: list[str] = []
    required_sheets = {"Ubicaciones", "Indicadores"}
    if not required_sheets.issubset(wb.sheetnames):
        missing = sorted(required_sheets.difference(wb.sheetnames))
        raise CMSValidationError([f"Faltan pestañas: {', '.join(missing)}"])

    locations = _read_sheet(wb["Ubicaciones"], LOCATION_HEADERS)
    metrics = _read_sheet(wb["Indicadores"], METRIC_HEADERS)

    location_by_cc: dict[str, dict[str, Any]] = {}
    for row in locations:
        cc = str(row.get("CC") or "").strip()
        label = f"Ubicaciones fila {row['_row']}"
        if not cc:
            errors.append(f"{label}: CC es obligatorio")
            continue
        if cc in location_by_cc:
            errors.append(f"{label}: CC duplicado {cc}")
            continue
        if not row.get("Nombre Tienda"):
            errors.append(f"{label}: Nombre Tienda es obligatorio")
        lat = _to_float(row.get("Latitud"), f"{label} Latitud", errors, False)
        lng = _to_float(row.get("Longitud"), f"{label} Longitud", errors, False)
        if lat is not None and not (-90 <= lat <= 90):
            errors.append(f"{label}: Latitud fuera de rango")
        if lng is not None and not (-180 <= lng <= 180):
            errors.append(f"{label}: Longitud fuera de rango")
        row["Latitud"], row["Longitud"] = lat, lng
        location_by_cc[cc] = row

    metric_by_cc: dict[str, dict[str, Any]] = {}
    for row in metrics:
        cc = str(row.get("CC") or "").strip()
        label = f"Indicadores fila {row['_row']}"
        if not cc:
            errors.append(f"{label}: CC es obligatorio")
            continue
        if cc in metric_by_cc:
            errors.append(f"{label}: CC duplicado {cc}")
            continue
        metric_by_cc[cc] = row

    missing_metrics = sorted(set(location_by_cc).difference(metric_by_cc))
    orphan_metrics = sorted(set(metric_by_cc).difference(location_by_cc))
    if missing_metrics:
        errors.append("CC sin indicadores: " + ", ".join(missing_metrics[:12]))
    if orphan_metrics:
        errors.append("Indicadores sin ubicación: " + ", ".join(orphan_metrics[:12]))

    stores = []
    for cc, location in location_by_cc.items():
        merged: dict[str, Any] = {}
        for header, field in LOCATION_MAP.items():
            merged[field] = location.get(header)
        metric = metric_by_cc.get(cc, {})
        for header, field in METRIC_MAP.items():
            if field != "cc":
                merged[field] = metric.get(header)
        merged["cc"] = cc
        active = merged.get("active")
        merged["active"] = str(active).strip().lower() not in {"no", "false", "0", "inactiva"}
        for field in NUMBER_FIELDS | PERCENT_FIELDS:
            if field in {"latitude", "longitude"}:
                continue
            merged[field] = _to_float(merged.get(field), f"CC {cc} {field}", errors)
        stores.append(merged)

    if errors:
        raise CMSValidationError(errors)

    stores.sort(key=lambda item: (str(item.get("dm") or ""), str(item.get("store_name") or "")))
    return {
        "metadata": {
            "schema_version": 2,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "source": "Excel CMS",
            "store_count": len(stores),
        },
        "stores": stores,
    }


def _backup(path: Path) -> None:
    if not path.exists():
        return
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    shutil.copy2(path, backup_dir / f"{path.stem}_{stamp}{path.suffix}")


def _atomic_write(path: Path, content: bytes, *, backup: bool = True) -> bool:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.read_bytes() == content:
        return False
    if backup:
        _backup(path)
    fd, temporary = tempfile.mkstemp(prefix=f"{path.stem}_", suffix=path.suffix, dir=path.parent)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)
    return True


def _json_bytes(payload: dict[str, Any]) -> bytes:
    return (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")


def _legacy_value(header: str, value: Any) -> str:
    if value is None:
        return "No hay valores" if header in {"DT Time", "Tiempo DT AA"} else ""
    if header in {"Fecha de Apertura", "Corte YTD"}:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%d/%m/%Y")
    if LEGACY_FIELDS[header] in PERCENT_FIELDS:
        return f"{float(value):.2%}"
    if header == "Ventas":
        return f"${float(value) / 1_000_000:.1f}M"
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def payload_to_csv_bytes(payload: dict[str, Any]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=LEGACY_HEADERS, lineterminator="\n")
    writer.writeheader()
    for store in sorted(payload["stores"], key=lambda item: str(item.get("cc") or "")):
        writer.writerow({header: _legacy_value(header, store.get(field)) for header, field in LEGACY_FIELDS.items()})
    return stream.getvalue().encode("utf-8-sig")


def _current_stores(database_path: Path) -> list[dict[str, Any]] | None:
    try:
        return json.loads(database_path.read_text(encoding="utf-8"))["stores"]
    except (OSError, ValueError, KeyError, TypeError):
        return None


def synchronize_outputs(
    workbook_path: str | Path,
    database_path: str | Path,
    legacy_csv_path: str | Path | None = None,
    *,
    dry_run: bool = False,
) -> tuple[dict[str, Any], list[str]]:
    """Sincroniza CMS→JSON/CSV; devuelve payload y archivos que cambian."""
    payload = workbook_to_payload(workbook_path)
    database_path = Path(database_path)
    changed: list[str] = []
    if _current_stores(database_path) != payload["stores"]:
        changed.append(str(database_path))
        if not dry_run:
            _atomic_write(database_path, _json_bytes(payload))
    if legacy_csv_path is not None:
        csv_path = Path(legacy_csv_path)
        csv_content = payload_to_csv_bytes(payload)
        if not csv_path.exists() or csv_path.read_bytes() != csv_content:
            changed.append(str(csv_path))
            if not dry_run:
                _atomic_write(csv_path, csv_content)
    return payload, changed


def apply_cms_workbook(
    workbook_path: str | Path,
    cms_path: str | Path,
    database_path: str | Path,
    legacy_csv_path: str | Path | None = None,
) -> tuple[dict[str, Any], list[str]]:
    """Instala el Excel validado y actualiza todos sus derivados."""
    workbook_to_payload(workbook_path)
    cms_path = Path(cms_path)
    changed: list[str] = []
    if _atomic_write(cms_path, Path(workbook_path).read_bytes()):
        changed.append(str(cms_path))
    payload, output_changes = synchronize_outputs(cms_path, database_path, legacy_csv_path)
    changed.extend(output_changes)
    return payload, changed


def update_database(workbook_path: str | Path, database_path: str | Path) -> dict[str, Any]:
    """Compatibilidad: actualiza solo JSON."""
    payload, _ = synchronize_outputs(workbook_path, database_path)
    return payload
